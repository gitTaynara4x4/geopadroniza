import os
import re
import time
import math
import uuid
import sqlite3
import threading
import unicodedata
from io import BytesIO
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional

import requests
from openpyxl import load_workbook
from fastapi import FastAPI, UploadFile, File, HTTPException, Response, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse, FileResponse

# =========================================================
# CONFIG
# =========================================================

BASE_DIR = Path(__file__).resolve().parent
APP_NAME = "GeoPadroniza"
APP_UA = "GeoPadroniza/SAFE-9.1 (+local-db-overwrite-no-geo+city-cep-fallback+smart-relocation+uppercase)"
APP_VERSION = "9.1.3"

NOMINATIM_DELAY_SECONDS = 1.1
REQUEST_TIMEOUT = 25

DNE_DB_PATH = os.getenv("GEOPADRONIZA_DNE_DB", str(BASE_DIR / "dne.db"))
MODES = {"ultra", "rapido", "completo"}

app = FastAPI(title=APP_NAME, version=APP_VERSION)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

JOBS: Dict[str, Dict[str, Any]] = {}

# =========================================================
# REGRAS
# =========================================================

PREPOSICOES = {"de", "da", "do", "das", "dos", "e"}

TIPOS_LOGRADOURO = [
    (r"^(avenida|av\.?)\b", "Av"),
    (r"^(rua|r\.?)\b", "R"),
    (r"^(estrada|estr\.?)\b", "Estr"),
    (r"^(travessa|tv\.?)\b", "Tv"),
    (r"^(pra[çc]a|p[çc]a\.?)\b", "Pça"),
    (r"^(largo|lgo\.?)\b", "Lgo"),
    (r"^(rodovia|rod\.?)\b", "Rod"),
    (r"^(alameda|al\.?)\b", "Al"),
]

CORRECOES_RUA = {
    r"\bd[\. ]?em[ií]lio winther\b": "Dr. Emílio Winther",
    r"\bd[\. ]?cesar costa\b": "Dr. César Costa",
    r"\bd[\. ]?cesacosta\b": "Dr. César Costa",
    r"\bd[\. ]?urbano figueira\b": "Dr. Urbano Figueira",
    r"\bjo[aã]o paulo de oliveira gama\b": "João Paulo de Oliveira Gama",
    r"\bvisconde do rio branco\b": "Visconde do Rio Branco",
    r"\banizio ortiz monteiro\b": "Anízio Ortiz Monteiro",
    r"\bitalia\b": "Itália",
    r"\bnove de julho\b": "Nove de Julho",
    r"\bantonio dias oliveira\b": "Antônio Dias Oliveira",
    r"\bpadre antonio diogo feij[oó]\b": "Padre Antônio Diogo Feijó",
}

TOKENS_RUINS = {
    "", "-", "--", "---", "n/a", "na", "null", "none",
    "sem", "nao informado", "não informado", "nao", "não",
    "teste", "x", "xx", "xxx", "0"
}

UF_BR = {
    "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO",
    "MA", "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI",
    "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO"
}

STATE_NAME_TO_UF = {
    "acre": "AC",
    "alagoas": "AL",
    "amapa": "AP",
    "amapá": "AP",
    "amazonas": "AM",
    "bahia": "BA",
    "ceara": "CE",
    "ceará": "CE",
    "distrito federal": "DF",
    "espirito santo": "ES",
    "espírito santo": "ES",
    "goias": "GO",
    "goiás": "GO",
    "maranhao": "MA",
    "maranhão": "MA",
    "mato grosso": "MT",
    "mato grosso do sul": "MS",
    "minas gerais": "MG",
    "para": "PA",
    "pará": "PA",
    "paraiba": "PB",
    "paraíba": "PB",
    "parana": "PR",
    "paraná": "PR",
    "pernambuco": "PE",
    "piaui": "PI",
    "piauí": "PI",
    "rio de janeiro": "RJ",
    "rio grande do norte": "RN",
    "rio grande do sul": "RS",
    "rondonia": "RO",
    "rondônia": "RO",
    "roraima": "RR",
    "santa catarina": "SC",
    "sao paulo": "SP",
    "são paulo": "SP",
    "sergipe": "SE",
    "tocantins": "TO",
}

HEADER_ALIASES = {
    "tipo": [
        "tipo logradouro", "tipo de logradouro", "tp logr", "tp logradouro", "tipo"
    ],
    "rua": [
        "logradouro", "rua", "endereco", "endereço", "logradouro cliente"
    ],
    "numero": [
        "numero", "número", "num", "n°", "nº", "num/end"
    ],
    "bairro": [
        "bairro", "bairro principal"
    ],
    "cidade": [
        "cidade", "municipio", "município", "municipio base"
    ],
    "uf": [
        "uf", "estado", "estado?"
    ],
    "cep": [
        "cep", "codigo postal", "código postal"
    ],
    "complemento": [
        "complemento", "compl", "comp1"
    ],
}

HEADER_SUFFIX_TOKENS = {
    "cliente", "principal", "entrega", "cobranca", "cobrança",
    "residencial", "comercial", "cadastro", "correspondencia",
    "correspondência", "base", "dados", "1", "2", "3"
}

TIPO_LOGRADOURO_TOKENS = {
    "avenida", "av", "rua", "r", "estrada", "estr", "travessa", "tv",
    "praca", "praça", "pca", "pça", "largo", "lgo", "rodovia", "rod",
    "alameda", "al"
}

TOKENS_GENERICOS_LOGRADOURO = {
    "avenida", "av", "rua", "r", "estrada", "estr", "travessa", "tv",
    "praca", "praça", "pca", "pça", "largo", "lgo", "rodovia", "rod",
    "alameda", "al", "doutor", "dr", "professor", "prof", "padre",
    "coronel", "cel", "visconde"
}

LOGRADOURO_HINTS = {
    "avenida", "av", "rua", "r", "estrada", "estr", "travessa", "tv",
    "praca", "praça", "pca", "pça", "largo", "lgo", "rodovia", "rod",
    "alameda", "al", "doutor", "dr", "professor", "prof", "padre",
    "visconde", "barao", "barão", "coronel", "cel",
    "oliveira", "figueira", "feijo", "feijó", "brasil",
    "winther", "costa", "novembro", "dutra"
}

BAIRRO_HINTS = {
    "centro", "jardim", "jd", "parque", "pq", "vila", "residencial",
    "chacara", "chácara", "bairro", "condominio", "condomínio",
    "loteamento", "distrito", "industrial", "esplanada", "olimpia", "olímpia"
}

COMPLEMENTO_HINTS = {
    "ap", "apto", "apartamento", "bl", "bloco", "fundos", "frente",
    "sala", "sl", "casa", "sobrado", "quadra", "qd", "lote", "lt",
    "andar", "loja", "box", "km"
}

ADMIN_TOKENS = {
    "ativo", "inativo", "pendente", "crm", "site", "email", "obs",
    "revisar", "origem", "status", "brasil", "teste"
}

# =========================================================
# UTIL
# =========================================================

def limpar_texto(valor):
    if valor is None:
        return ""
    if isinstance(valor, float) and math.isnan(valor):
        return ""
    texto = str(valor).strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def remover_acentos(texto):
    texto = limpar_texto(texto)
    return "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )


def normalizar_chave(texto):
    texto = remover_acentos(texto).lower()
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def tokenizar_normalizado(texto: str) -> List[str]:
    texto = normalizar_chave(texto)
    return [t for t in re.split(r"[^a-z0-9]+", texto) if t]


def smart_title(texto):
    texto = limpar_texto(texto)
    if not texto:
        return ""
    partes = texto.lower().split(" ")
    out = []
    for i, p in enumerate(partes):
        if i > 0 and p in PREPOSICOES:
            out.append(p)
        else:
            out.append(p[:1].upper() + p[1:])
    return " ".join(out)


def so_digitos(texto):
    return re.sub(r"\D", "", limpar_texto(texto))


def form_bool(valor) -> bool:
    return str(valor).strip().lower() in {"1", "true", "on", "yes", "sim"}


def texto_saida_endereco(valor: str, caixa_alta: bool = False) -> str:
    txt = limpar_texto(valor)
    if not txt:
        return ""
    return txt.upper() if caixa_alta else txt


def cep_eh_placeholder(cep: str) -> bool:
    c = so_digitos(cep)
    if not c:
        return True
    if c in {"0", "00000000"}:
        return True
    if len(c) != 8:
        return True
    if len(set(c)) == 1:
        return True
    return False


def extrair_cep_de_texto(texto: str) -> str:
    txt = limpar_texto(texto)
    if not txt:
        return ""
    m = re.search(r"\b(\d{5})[-\s]?(\d{3})\b", txt)
    if not m:
        return ""
    cep = f"{m.group(1)}{m.group(2)}"
    return formatar_cep(cep) if not cep_eh_placeholder(cep) else ""


def remover_cep_do_texto(texto: str) -> str:
    txt = limpar_texto(texto)
    if not txt:
        return ""
    txt = re.sub(r"\bcep\b[:\s-]*\d{5}[-\s]?\d{3}\b", "", txt, flags=re.IGNORECASE)
    txt = re.sub(r"\b\d{5}[-\s]?\d{3}\b", "", txt)
    txt = re.sub(r"\s+", " ", txt).strip(" ,;/-")
    return txt


def texto_tem_padrao_cep(texto: str) -> bool:
    return bool(extrair_cep_de_texto(texto))


def texto_tem_digitos_ruins_localidade(texto: str) -> bool:
    t = limpar_texto(texto)
    if not t:
        return False
    if texto_tem_padrao_cep(t):
        return True
    digitos = re.findall(r"\d", t)
    if len(digitos) >= 2:
        return True
    if re.search(r"\d+\s*/\s*\d+", t):
        return True
    return False


def formatar_cep(cep):
    c = so_digitos(cep)
    if len(c) == 8 and not cep_eh_placeholder(c):
        return f"{c[:5]}-{c[5:]}"
    return ""


def normalizar_uf(uf):
    t = limpar_texto(uf)
    if not t:
        return ""
    n = normalizar_chave(t)
    compacto = re.sub(r"[^a-z]", "", n)

    if len(compacto) == 2 and compacto.upper() in UF_BR:
        return compacto.upper()

    if n in STATE_NAME_TO_UF:
        return STATE_NAME_TO_UF[n]

    return ""


def uf_valida(uf):
    return normalizar_uf(uf) in UF_BR


def normalizar_numero(numero):
    n = normalizar_chave(numero)
    if not n:
        return ""
    if n in {"s/n", "sn", "sem numero", "sem número"}:
        return "s/n"

    m = re.search(r"\b(\d{1,6}[a-z]?)\b", n)
    if m:
        valor = m.group(1).upper()
        if re.fullmatch(r"0+", so_digitos(valor)):
            return ""
        return valor

    return ""


def numero_valido(numero):
    n = normalizar_chave(numero)
    if not n:
        return False
    if n in {"s/n", "sn"}:
        return True
    if re.fullmatch(r"0+", so_digitos(n)):
        return False
    return bool(re.search(r"\d", n))


def rua_suspeita(rua):
    r = normalizar_chave(rua)
    if not r or r in TOKENS_RUINS:
        return True
    if len(r) < 4:
        return True
    if texto_tem_padrao_cep(r):
        return True
    if sum(ch.isdigit() for ch in r) > len(r) * 0.4:
        return True
    return False


def bairro_suspeito(bairro):
    b = normalizar_chave(bairro)
    if not b or b in TOKENS_RUINS or len(b) < 3:
        return True
    if b in STATE_NAME_TO_UF:
        return True
    if texto_tem_digitos_ruins_localidade(bairro):
        return True
    return False


def cidade_suspeita(cidade):
    c = normalizar_chave(cidade)
    if not c or c in TOKENS_RUINS or len(c) < 3:
        return True
    if c in STATE_NAME_TO_UF:
        return True
    if texto_tem_digitos_ruins_localidade(cidade):
        return True
    toks = set(tokenizar_normalizado(c))
    if toks & BAIRRO_HINTS:
        return True
    return False


def cep_valido(cep):
    c = so_digitos(cep)
    return len(c) == 8 and not cep_eh_placeholder(c)


def render_tipo_logradouro(texto):
    t = limpar_texto(texto)
    if not t:
        return ""
    tn = normalizar_chave(t)
    for padrao, novo in TIPOS_LOGRADOURO:
        if re.match(padrao, tn, flags=re.IGNORECASE):
            return novo
    return smart_title(t)


def corrigir_rua(rua):
    r = limpar_texto(rua)
    if not r:
        return ""
    for padrao, correcao in CORRECOES_RUA.items():
        r = re.sub(padrao, correcao, r, flags=re.IGNORECASE)
    r = smart_title(r)
    r = re.sub(r"\bDr\b\.?", "Dr.", r)
    r = re.sub(r"\bPca\b", "Pça", r)
    return r


def separar_tipo_nome_logradouro(logradouro):
    txt = limpar_texto(logradouro)
    if not txt:
        return "", ""
    txt_norm = normalizar_chave(txt)
    for padrao, tipo in TIPOS_LOGRADOURO:
        if re.match(padrao, txt_norm, flags=re.IGNORECASE):
            partes = txt.split(" ", 1)
            nome = partes[1] if len(partes) > 1 else ""
            return tipo, smart_title(nome)
    return "", smart_title(txt)


def extrair_numero_do_logradouro(logradouro):
    txt = limpar_texto(logradouro)
    if not txt:
        return "", ""

    m = re.search(r"^(.*?)[,\s\-]+(?:n[º°o]?\s*)?(\d{1,6}[A-Za-z]?)$", txt)
    if m:
        base = m.group(1).strip(" ,-")
        num = m.group(2).upper()
        if base:
            return base, num

    return txt, ""


def montar_logradouro(tipo, rua):
    tipo = limpar_texto(tipo)
    rua = limpar_texto(rua)
    if tipo and rua:
        rua_norm = normalizar_chave(rua)
        tipo_norm = normalizar_chave(tipo)
        if rua_norm.startswith(tipo_norm + " "):
            return rua
        return f"{tipo} {rua}".strip()
    return rua or tipo


def normalizar_localidade(texto):
    t = limpar_texto(texto)
    if not t:
        return ""
    t = remover_cep_do_texto(t)
    n = normalizar_chave(t)
    if n in TOKENS_RUINS:
        return ""
    if n in ADMIN_TOKENS:
        return ""
    if n in STATE_NAME_TO_UF:
        return ""
    if normalizar_uf(t):
        return ""
    if texto_tem_digitos_ruins_localidade(t):
        return ""
    if len(n) < 3:
        return ""
    return smart_title(t)


def montar_endereco_consulta(campos):
    via = montar_logradouro(campos.get("tipo", ""), campos.get("rua", ""))
    partes = [
        via,
        campos.get("numero", ""),
        campos.get("complemento", ""),
        campos.get("bairro", ""),
        campos.get("cidade", ""),
        campos.get("uf", ""),
        campos.get("cep", ""),
        "Brasil",
    ]
    partes = [limpar_texto(x) for x in partes if limpar_texto(x)]
    return ", ".join(partes)


def similaridade(a, b):
    a = normalizar_chave(a)
    b = normalizar_chave(b)
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    common = len(set(a.split()) & set(b.split()))
    return common / max(len(set(a.split()) | set(b.split())), 1)


def token_logradouro_para_busca(logradouro: str) -> str:
    toks = [
        t for t in tokenizar_normalizado(logradouro)
        if t not in TOKENS_GENERICOS_LOGRADOURO
        and t not in PREPOSICOES
        and len(t) >= 4
    ]
    if toks:
        toks = sorted(toks, key=len, reverse=True)
        return toks[0]
    base = normalizar_chave(logradouro)
    return base[:12] if base else ""


def cep_original_invalido_e_nao_aproveitavel(valor: str) -> bool:
    v = limpar_texto(valor)
    if not v:
        return False
    return not cep_valido(v)


def cidade_original_invalida_e_nao_aproveitavel(valor: str) -> bool:
    v = limpar_texto(valor)
    if not v:
        return False
    return cidade_suspeita(v)


def bairro_original_invalido_e_nao_aproveitavel(valor: str) -> bool:
    v = limpar_texto(valor)
    if not v:
        return False
    return bairro_suspeito(v)


def uf_original_invalida_e_nao_aproveitavel(valor: str) -> bool:
    v = limpar_texto(valor)
    if not v:
        return False
    return not uf_valida(v)

# =========================================================
# JOBS
# =========================================================

def init_job(filename: str, mode: str) -> str:
    job_id = uuid.uuid4().hex
    JOBS[job_id] = {
        "id": job_id,
        "filename": filename,
        "mode": mode,
        "status": "aguardando",
        "progress": 0,
        "current": 0,
        "total": 0,
        "message": "Aguardando processamento...",
        "done": False,
        "error": None,
        "output_bytes": None,
        "output_name": None,
        "uppercase_addresses": False,
    }
    return job_id


def update_job(job_id: str, **kwargs):
    if job_id in JOBS:
        JOBS[job_id].update(kwargs)

# =========================================================
# CONTEXTO
# =========================================================

class ConsultaContexto:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": APP_UA,
            "Accept": "application/json",
        })
        self.cache_viacep_cep = {}
        self.cache_viacep_busca = {}
        self.cache_nominatim = {}
        self.cache_dne_cep = {}
        self.cache_dne_endereco = {}
        self.cache_dne_cidade = {}
        self._last_nominatim = 0.0
        self.db_path = DNE_DB_PATH
        self.has_local_db = Path(self.db_path).exists()
        self._conn = None

    def aguardar_nominatim(self):
        agora = time.time()
        delta = agora - self._last_nominatim
        if delta < NOMINATIM_DELAY_SECONDS:
            time.sleep(NOMINATIM_DELAY_SECONDS - delta)
        self._last_nominatim = time.time()

    def get_db(self):
        if not self.has_local_db:
            return None
        if self._conn is None:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            conn.create_function("NORMTXT", 1, lambda v: normalizar_chave(v or ""))
            self._conn = conn
        return self._conn

    def close(self):
        try:
            if self._conn is not None:
                self._conn.close()
        except Exception:
            pass
        self._conn = None

# =========================================================
# BASE LOCAL CORREIOS
# =========================================================

def _row_para_dict_correios(row: sqlite3.Row):
    if not row:
        return None
    return {
        "cep": limpar_texto(row["cep"]),
        "logradouro": limpar_texto(row["logradouro"] or row["nome"]),
        "bairro": limpar_texto(row["bairro"]),
        "localidade": limpar_texto(row["municipio"]),
        "uf": limpar_texto(row["uf"]),
        "complemento": limpar_texto(row["complemento"]),
        "nome": limpar_texto(row["nome"]),
    }


def correios_por_cep_local(cep: str, ctx: ConsultaContexto):
    cep_limpo = so_digitos(cep)
    if len(cep_limpo) != 8 or cep_eh_placeholder(cep_limpo):
        return None

    if cep_limpo in ctx.cache_dne_cep:
        return ctx.cache_dne_cep[cep_limpo]

    conn = ctx.get_db()
    if conn is None:
        ctx.cache_dne_cep[cep_limpo] = None
        return None

    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT cep, logradouro, complemento, bairro, municipio, uf, nome
            FROM cep_unificado
            WHERE cep = ?
            LIMIT 1
            """,
            (cep_limpo,),
        )
        row = cur.fetchone()
        data = _row_para_dict_correios(row)
        ctx.cache_dne_cep[cep_limpo] = data
        return data
    except Exception:
        ctx.cache_dne_cep[cep_limpo] = None
        return None


def correios_por_endereco_local(uf: str, cidade: str, logradouro: str, bairro: str, ctx: ConsultaContexto):
    uf_n = normalizar_uf(uf)
    cidade_n = normalizar_chave(cidade)
    logradouro_n = normalizar_chave(logradouro)
    bairro_n = normalizar_chave(bairro)

    if not uf_n or len(cidade_n) < 3 or len(logradouro_n) < 4:
        return None

    termo = token_logradouro_para_busca(logradouro)
    if len(termo) < 3:
        return None

    chave = (uf_n, cidade_n, logradouro_n, bairro_n)
    if chave in ctx.cache_dne_endereco:
        return ctx.cache_dne_endereco[chave]

    conn = ctx.get_db()
    if conn is None:
        ctx.cache_dne_endereco[chave] = None
        return None

    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT cep, logradouro, complemento, bairro, municipio, uf, nome
            FROM cep_unificado
            WHERE uf = ?
              AND NORMTXT(municipio) = ?
              AND (
                    NORMTXT(logradouro) LIKE ?
                    OR NORMTXT(nome) LIKE ?
                  )
            LIMIT 120
            """,
            (uf_n, cidade_n, f"%{termo}%", f"%{termo}%"),
        )
        rows = cur.fetchall()

        melhor = None
        melhor_score = -1.0

        for row in rows:
            item = _row_para_dict_correios(row)
            score = similaridade(logradouro, item.get("logradouro", ""))
            if bairro:
                score += similaridade(bairro, item.get("bairro", "")) * 0.35
            if cidade_n == normalizar_chave(item.get("localidade", "")):
                score += 0.20

            if melhor is None or score > melhor_score:
                melhor = item
                melhor_score = score

        if melhor_score < 0.45:
            melhor = None

        ctx.cache_dne_endereco[chave] = melhor
        return melhor
    except Exception:
        ctx.cache_dne_endereco[chave] = None
        return None


def correios_aproximado_por_cidade_local(uf: str, cidade: str, ctx: ConsultaContexto):
    uf_n = normalizar_uf(uf)
    cidade_n = normalizar_chave(cidade)
    if not uf_n or len(cidade_n) < 3:
        return None

    chave = (uf_n, cidade_n)
    if chave in ctx.cache_dne_cidade:
        return ctx.cache_dne_cidade[chave]

    conn = ctx.get_db()
    if conn is None:
        ctx.cache_dne_cidade[chave] = None
        return None

    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT cep, logradouro, complemento, bairro, municipio, uf, nome
            FROM cep_unificado
            WHERE uf = ?
              AND NORMTXT(municipio) = ?
            LIMIT 400
            """,
            (uf_n, cidade_n),
        )
        rows = cur.fetchall()
        if not rows:
            ctx.cache_dne_cidade[chave] = None
            return None

        candidatos = []
        for row in rows:
            item = _row_para_dict_correios(row)
            score = 0.0
            bairro_n = normalizar_chave(item.get("bairro", ""))
            logr_n = normalizar_chave(item.get("logradouro", ""))
            nome_n = normalizar_chave(item.get("nome", ""))
            if "centro" in bairro_n:
                score += 2.0
            if "centro" in logr_n or "centro" in nome_n:
                score += 0.5
            if item.get("cep"):
                score += 0.2
            candidatos.append((score, so_digitos(item.get("cep", "")), item))

        candidatos.sort(key=lambda x: (-x[0], x[1] or "99999999"))
        melhor = candidatos[0][2] if candidatos else None
        ctx.cache_dne_cidade[chave] = melhor
        return melhor
    except Exception:
        ctx.cache_dne_cidade[chave] = None
        return None

# =========================================================
# VIACEP
# =========================================================

def viacep_por_cep(cep, ctx: ConsultaContexto):
    cep_limpo = so_digitos(cep)
    if len(cep_limpo) != 8 or cep_eh_placeholder(cep_limpo):
        return None

    if cep_limpo in ctx.cache_viacep_cep:
        return ctx.cache_viacep_cep[cep_limpo]

    url = f"https://viacep.com.br/ws/{cep_limpo}/json/"
    try:
        resp = ctx.session.get(url, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        data = resp.json()
        if data.get("erro"):
            ctx.cache_viacep_cep[cep_limpo] = None
            return None
        ctx.cache_viacep_cep[cep_limpo] = data
        return data
    except Exception:
        ctx.cache_viacep_cep[cep_limpo] = None
        return None


def escolher_melhor_resultado_viacep(resultados, logradouro_alvo="", bairro_alvo=""):
    if not isinstance(resultados, list) or not resultados:
        return None

    melhor = None
    melhor_score = -1.0

    for item in resultados:
        s1 = similaridade(logradouro_alvo, item.get("logradouro", ""))
        s2 = similaridade(bairro_alvo, item.get("bairro", "")) if bairro_alvo else 0
        score = s1 + (s2 * 0.35)
        if melhor is None or score > melhor_score:
            melhor = item
            melhor_score = score

    return melhor


def viacep_por_endereco(uf, cidade, logradouro, bairro, ctx: ConsultaContexto):
    uf = normalizar_uf(uf)
    cidade = limpar_texto(cidade)
    logradouro = limpar_texto(logradouro)

    if len(uf) != 2 or len(cidade) < 2 or len(logradouro) < 3:
        return None

    chave = (uf, normalizar_chave(cidade), normalizar_chave(logradouro))
    if chave in ctx.cache_viacep_busca:
        return ctx.cache_viacep_busca[chave]

    cidade_url = requests.utils.quote(cidade)
    logradouro_url = requests.utils.quote(logradouro)
    url = f"https://viacep.com.br/ws/{uf}/{cidade_url}/{logradouro_url}/json/"

    try:
        resp = ctx.session.get(url, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        data = resp.json()
        melhor = escolher_melhor_resultado_viacep(data, logradouro, bairro)
        ctx.cache_viacep_busca[chave] = melhor
        return melhor
    except Exception:
        ctx.cache_viacep_busca[chave] = None
        return None

# =========================================================
# OSM
# =========================================================

def nominatim_buscar(endereco, ctx: ConsultaContexto):
    q = limpar_texto(endereco)
    if not q:
        return None

    if q in ctx.cache_nominatim:
        return ctx.cache_nominatim[q]

    try:
        ctx.aguardar_nominatim()
        resp = ctx.session.get(
            "https://nominatim.openstreetmap.org/search",
            params={
                "q": q,
                "format": "jsonv2",
                "addressdetails": 1,
                "limit": 1,
                "countrycodes": "br",
                "accept-language": "pt-BR",
            },
            timeout=REQUEST_TIMEOUT,
        )
        resp.raise_for_status()
        data = resp.json()
        item = data[0] if data else None
        ctx.cache_nominatim[q] = item
        return item
    except Exception:
        ctx.cache_nominatim[q] = None
        return None


def pegar_osm_address(item, *keys):
    if not item:
        return ""
    addr = item.get("address", {}) or {}
    for k in keys:
        if addr.get(k):
            return addr.get(k)
    return ""

# =========================================================
# DETECÇÃO DE COLUNAS
# =========================================================

def nome_coluna_combina_alias(nome_coluna: str, alias: str) -> bool:
    nome = normalizar_chave(nome_coluna)
    alias = normalizar_chave(alias)

    if not nome or not alias:
        return False

    if nome == alias:
        return True

    if nome.startswith(alias + " "):
        resto = nome[len(alias):].strip()
        resto_tokens = set(tokenizar_normalizado(resto))
        if resto_tokens and resto_tokens.issubset(HEADER_SUFFIX_TOKENS):
            return True

    if nome.endswith(" " + alias):
        prefixo = nome[:-len(alias)].strip()
        prefixo_tokens = set(tokenizar_normalizado(prefixo))
        if prefixo_tokens and prefixo_tokens.issubset({"campo", "dados", "info", "informacao", "informação"}):
            return True

    return False


def header_map(ws):
    cols = {}
    for c in range(1, ws.max_column + 1):
        cols[c] = normalizar_chave(ws.cell(row=1, column=c).value)
    return cols


def achar_coluna_por_header(headers: Dict[int, str], aliases: List[str]):
    aliases_n = [normalizar_chave(a) for a in aliases]

    for alias in aliases_n:
        for idx, nome in headers.items():
            if nome == alias:
                return idx

    for alias in aliases_n:
        for idx, nome in headers.items():
            if nome_coluna_combina_alias(nome, alias):
                return idx

    return None


def detectar_header_na_primeira_linha(ws):
    headers = header_map(ws)

    cols = {
        "tipo": achar_coluna_por_header(headers, HEADER_ALIASES["tipo"]),
        "rua": achar_coluna_por_header(headers, HEADER_ALIASES["rua"]),
        "numero": achar_coluna_por_header(headers, HEADER_ALIASES["numero"]),
        "bairro": achar_coluna_por_header(headers, HEADER_ALIASES["bairro"]),
        "cidade": achar_coluna_por_header(headers, HEADER_ALIASES["cidade"]),
        "uf": achar_coluna_por_header(headers, HEADER_ALIASES["uf"]),
        "cep": achar_coluna_por_header(headers, HEADER_ALIASES["cep"]),
        "complemento": achar_coluna_por_header(headers, HEADER_ALIASES["complemento"]),
    }

    hits = sum(1 for v in cols.values() if v)
    tem_campo_chave = bool(cols["rua"] or cols["cep"] or cols["cidade"] or cols["bairro"])

    if hits >= 2 and tem_campo_chave:
        return {
            "source": "header_linha_1",
            "start_row": 2,
            "cols": cols,
        }

    return None


def amostras_coluna(ws, col_idx: int, start_row: int = 1, max_rows: int = 30) -> List[str]:
    valores = []
    fim = min(ws.max_row, start_row + max_rows - 1)
    for r in range(start_row, fim + 1):
        v = limpar_texto(ws.cell(row=r, column=col_idx).value)
        if v:
            valores.append(v)
    return valores


def ratio_true(fn, valores: List[str]) -> float:
    if not valores:
        return 0.0
    return sum(1 for v in valores if fn(v)) / len(valores)


def parece_cep_valor(valor: str) -> bool:
    return cep_valido(valor)


def parece_uf_valor(valor: str) -> bool:
    return bool(normalizar_uf(valor))


def parece_tipo_logradouro_valor(valor: str) -> bool:
    v = normalizar_chave(valor).replace(".", "")
    return v in TIPO_LOGRADOURO_TOKENS


def parece_numero_valor(valor: str) -> bool:
    v = normalizar_chave(valor)
    if not v:
        return False
    if v in {"s/n", "sn", "sem numero", "sem número"}:
        return True
    if parece_cep_valor(v):
        return False
    if re.fullmatch(r"0+", so_digitos(v)):
        return False
    return bool(re.fullmatch(r"(n[º°o]?\s*)?\d{1,6}[a-z]?", v))


def parece_complemento_valor(valor: str) -> bool:
    toks = set(tokenizar_normalizado(valor))
    if not toks:
        return False
    if toks & COMPLEMENTO_HINTS:
        return True
    v = normalizar_chave(valor)
    return bool(re.search(r"\b(apto|apartamento|bloco|sala|fundos|quadra|lote|andar|loja|km)\b", v))


def parece_texto_localidade(valor: str) -> bool:
    v = limpar_texto(valor)
    if not v:
        return False
    if parece_cep_valor(v) or parece_uf_valor(v) or parece_numero_valor(v):
        return False

    vn = normalizar_chave(v)
    if vn in TOKENS_RUINS:
        return False
    if vn in ADMIN_TOKENS:
        return False
    if vn in STATE_NAME_TO_UF:
        return False
    if len(vn) < 3:
        return False
    if texto_tem_digitos_ruins_localidade(v):
        return False
    return True


def parece_bairro_valor(valor: str) -> bool:
    if not parece_texto_localidade(valor):
        return False
    toks = set(tokenizar_normalizado(valor))
    if toks & BAIRRO_HINTS:
        return True
    return 1 <= len(toks) <= 5


def parece_cidade_valor(valor: str) -> bool:
    if not parece_texto_localidade(valor):
        return False
    toks = set(tokenizar_normalizado(valor))
    if toks & BAIRRO_HINTS:
        return False
    return 1 <= len(toks) <= 4


def parece_logradouro_valor(valor: str) -> bool:
    if not valor:
        return False

    v = normalizar_chave(valor)
    if v in TOKENS_RUINS:
        return False
    if parece_cep_valor(v) or parece_uf_valor(v):
        return False

    if re.match(
        r"^(avenida|av\.?|rua|r\.?|estrada|estr\.?|travessa|tv\.?|pra[çc]a|p[çc]a\.?|largo|lgo\.?|rodovia|rod\.?|alameda|al\.?|doutor|dr\.?|professor|prof\.?|padre|visconde|coronel|cel\.?)\b",
        v,
        flags=re.IGNORECASE,
    ):
        return True

    toks = set(tokenizar_normalizado(v))
    if toks & LOGRADOURO_HINTS and len(toks) >= 2:
        return True

    return False


def score_coluna_admin(valores: List[str]) -> float:
    if not valores:
        return 0.0
    score = 0.0
    for v in valores:
        n = normalizar_chave(v)
        toks = set(tokenizar_normalizado(v))
        if n in ADMIN_TOKENS:
            score += 2.0
        if toks and toks.issubset({"l1", "l2", "a", "b", "c", "obs", "crm", "site", "email", "pendente", "ativo", "inativo", "revisar"}):
            score += 2.0
    return score / len(valores)


def pontuar_coluna_por_amostras(ws, col_idx: int) -> Dict[str, float]:
    valores = amostras_coluna(ws, col_idx, start_row=1, max_rows=30)
    scores = {
        "tipo": 0.0,
        "rua": 0.0,
        "numero": 0.0,
        "bairro": 0.0,
        "cidade": 0.0,
        "uf": 0.0,
        "cep": 0.0,
        "complemento": 0.0,
        "admin": score_coluna_admin(valores),
    }

    if not valores:
        return scores

    for v in valores:
        if parece_cep_valor(v):
            scores["cep"] += 5.0
            continue

        if parece_uf_valor(v):
            scores["uf"] += 5.0
            continue

        if parece_tipo_logradouro_valor(v):
            scores["tipo"] += 4.0

        if parece_numero_valor(v):
            scores["numero"] += 3.4

        if parece_complemento_valor(v):
            scores["complemento"] += 3.0

        if parece_logradouro_valor(v):
            scores["rua"] += 4.2

        if parece_texto_localidade(v):
            scores["bairro"] += 1.0
            scores["cidade"] += 1.0

            if parece_bairro_valor(v):
                scores["bairro"] += 1.7

            if parece_cidade_valor(v):
                scores["cidade"] += 1.7

    return scores


def escolher_melhor_coluna(score_map: Dict[int, Dict[str, float]], field: str, used: set, min_score: float):
    melhor_col = None
    melhor_score = min_score

    for col_idx, scores in score_map.items():
        if col_idx in used:
            continue
        if scores.get("admin", 0.0) >= 1.2:
            continue

        score = scores.get(field, 0.0)
        if score > melhor_score:
            melhor_score = score
            melhor_col = col_idx

    return melhor_col, melhor_score


def mapeamento_tem_sinais(cols: Dict[str, Optional[int]]) -> bool:
    principais = sum(
        1 for k in ["rua", "bairro", "cidade", "cep", "uf", "numero", "tipo"]
        if cols.get(k)
    )

    tem_estrutura = bool(
        cols.get("rua") and (cols.get("bairro") or cols.get("cidade") or cols.get("cep") or cols.get("uf"))
    )

    tem_cep_estrutura = bool(
        cols.get("cep") and (cols.get("rua") or cols.get("cidade") or cols.get("bairro"))
    )

    return principais >= 3 and (tem_estrutura or tem_cep_estrutura)


def detectar_mapeamento_por_inferencia(ws):
    score_map = {}
    for c in range(1, ws.max_column + 1):
        score_map[c] = pontuar_coluna_por_amostras(ws, c)

    used = set()
    cols = {
        "tipo": None,
        "rua": None,
        "numero": None,
        "bairro": None,
        "cidade": None,
        "uf": None,
        "cep": None,
        "complemento": None,
    }

    for field, min_score in [
        ("cep", 4.0),
        ("uf", 4.0),
        ("rua", 3.5),
        ("tipo", 3.0),
        ("numero", 2.8),
        ("complemento", 2.6),
        ("bairro", 2.1),
        ("cidade", 2.1),
    ]:
        col, _ = escolher_melhor_coluna(score_map, field, used, min_score)
        if col:
            cols[field] = col
            used.add(col)

    if not mapeamento_tem_sinais(cols):
        return None

    return {
        "source": "inferencia_inteligente_linha_1",
        "start_row": 1,
        "cols": cols,
    }


def validar_mapeamento_inferido(ws, det: Dict[str, Any]) -> bool:
    cols = det["cols"]

    rua_ok = False
    cidade_ok = False
    uf_ok = False
    cep_ok = False

    if cols.get("rua"):
        vals = amostras_coluna(ws, cols["rua"], 1, 25)
        rua_ok = ratio_true(parece_logradouro_valor, vals) >= 0.35

    if cols.get("cidade"):
        vals = amostras_coluna(ws, cols["cidade"], 1, 25)
        cidade_ok = ratio_true(parece_cidade_valor, vals) >= 0.25

    if cols.get("uf"):
        vals = amostras_coluna(ws, cols["uf"], 1, 25)
        uf_ok = ratio_true(parece_uf_valor, vals) >= 0.25

    if cols.get("cep"):
        vals = amostras_coluna(ws, cols["cep"], 1, 25)
        cep_ok = ratio_true(parece_cep_valor, vals) >= 0.15

    if not (rua_ok or cep_ok):
        return False

    secundarios = sum([cidade_ok, uf_ok, cep_ok])
    return secundarios >= 1


def detectar_mapeamento_ws(ws):
    header_det = detectar_header_na_primeira_linha(ws)
    if header_det:
        return header_det

    infer_det = detectar_mapeamento_por_inferencia(ws)
    if infer_det and validar_mapeamento_inferido(ws, infer_det):
        return infer_det

    return {
        "source": None,
        "start_row": 1,
        "cols": {
            "tipo": None,
            "rua": None,
            "numero": None,
            "bairro": None,
            "cidade": None,
            "uf": None,
            "cep": None,
            "complemento": None,
        }
    }

# =========================================================
# LIMPEZA DE ARTEFATOS ANTIGOS
# =========================================================

def remover_colunas_geo(ws):
    cols_geo = []
    for c in range(1, ws.max_column + 1):
        nome = limpar_texto(ws.cell(row=1, column=c).value)
        if nome.lower().startswith("geo_"):
            cols_geo.append(c)

    for col_idx in reversed(cols_geo):
        ws.delete_cols(col_idx, 1)


def limpar_artefatos_antigos(wb):
    if "LOG_PROCESSAMENTO" in wb.sheetnames:
        del wb["LOG_PROCESSAMENTO"]

    for ws in wb.worksheets:
        remover_colunas_geo(ws)

# =========================================================
# LEITURA / LIMPEZA DE LINHA
# =========================================================

def parece_bairro_texto(texto: str) -> bool:
    n = normalizar_chave(texto)
    if not n:
        return False
    toks = set(tokenizar_normalizado(n))
    return bool(toks & BAIRRO_HINTS)


def parece_cidade_texto(texto: str) -> bool:
    n = normalizar_chave(texto)
    if not n:
        return False
    if n in STATE_NAME_TO_UF:
        return False
    if texto_tem_digitos_ruins_localidade(texto):
        return False
    toks = set(tokenizar_normalizado(n))
    if toks & BAIRRO_HINTS:
        return False
    return len(toks) <= 4


def aplicar_realocacao_inteligente(campos: Dict[str, str]) -> Dict[str, str]:
    campos = {k: limpar_texto(v) for k, v in campos.items()}

    def mover(origem: str, destino: str, valor: str):
        valor = limpar_texto(valor)
        if not valor:
            campos[origem] = ""
            return
        if not limpar_texto(campos.get(destino, "")):
            campos[destino] = valor
        campos[origem] = ""

    cidade = limpar_texto(campos.get("cidade", ""))
    if cidade:
        if normalizar_uf(cidade):
            mover("cidade", "uf", normalizar_uf(cidade))
        elif extrair_cep_de_texto(cidade):
            mover("cidade", "cep", extrair_cep_de_texto(cidade))
        elif parece_bairro_texto(cidade) and not parece_cidade_texto(cidade):
            mover("cidade", "bairro", cidade)

    bairro = limpar_texto(campos.get("bairro", ""))
    if bairro:
        if normalizar_uf(bairro):
            mover("bairro", "uf", normalizar_uf(bairro))
        elif extrair_cep_de_texto(bairro):
            mover("bairro", "cep", extrair_cep_de_texto(bairro))
        elif parece_cidade_texto(bairro) and not parece_bairro_texto(bairro):
            mover("bairro", "cidade", bairro)

    cep = limpar_texto(campos.get("cep", ""))
    if cep and not cep_valido(cep):
        if normalizar_uf(cep):
            mover("cep", "uf", normalizar_uf(cep))
        elif parece_bairro_texto(cep):
            mover("cep", "bairro", cep)
        elif parece_cidade_texto(cep):
            mover("cep", "cidade", cep)

    uf = limpar_texto(campos.get("uf", ""))
    if uf and not uf_valida(uf):
        if extrair_cep_de_texto(uf):
            mover("uf", "cep", extrair_cep_de_texto(uf))
        elif parece_bairro_texto(uf):
            mover("uf", "bairro", uf)
        elif parece_cidade_texto(uf):
            mover("uf", "cidade", uf)

    complemento = limpar_texto(campos.get("complemento", ""))
    if complemento:
        if not campos.get("cep") and extrair_cep_de_texto(complemento):
            campos["cep"] = extrair_cep_de_texto(complemento)
            campos["complemento"] = remover_cep_do_texto(complemento)
        complemento = limpar_texto(campos.get("complemento", ""))
        if complemento and not campos.get("uf") and normalizar_uf(complemento):
            campos["uf"] = normalizar_uf(complemento)
            campos["complemento"] = ""
        elif complemento and not campos.get("bairro") and parece_bairro_texto(complemento):
            campos["bairro"] = complemento
            campos["complemento"] = ""
        elif complemento and not campos.get("cidade") and parece_cidade_texto(complemento):
            campos["cidade"] = complemento
            campos["complemento"] = ""

    return campos


def ler_campos_da_linha(ws, row_idx: int, cols_map: Dict[str, Any]):
    def v(col_idx):
        if not col_idx:
            return ""
        return limpar_texto(ws.cell(row=row_idx, column=col_idx).value)

    tipo_raw = v(cols_map["tipo"])
    rua_raw = v(cols_map["rua"])
    numero_raw = v(cols_map["numero"])
    bairro_raw = v(cols_map["bairro"])
    cidade_raw = v(cols_map["cidade"])
    uf_raw = v(cols_map["uf"])
    cep_raw = v(cols_map["cep"])
    complemento_raw = v(cols_map["complemento"])

    cep_embutido = (
        extrair_cep_de_texto(cep_raw)
        or extrair_cep_de_texto(cidade_raw)
        or extrair_cep_de_texto(bairro_raw)
        or extrair_cep_de_texto(complemento_raw)
    )

    cidade_raw = remover_cep_do_texto(cidade_raw)
    bairro_raw = remover_cep_do_texto(bairro_raw)
    complemento_raw = remover_cep_do_texto(complemento_raw)

    rua_sem_num, numero_embutido = extrair_numero_do_logradouro(rua_raw)
    tipo_emb, nome_emb = separar_tipo_nome_logradouro(rua_sem_num)

    tipo = render_tipo_logradouro(tipo_raw)

    if tipo_emb:
        tipo = tipo_emb
        rua = nome_emb
    else:
        rua = rua_sem_num

    rua = corrigir_rua(rua)

    numero = normalizar_numero(numero_raw)
    if not numero:
        numero = normalizar_numero(numero_embutido)

    if not numero and parece_numero_valor(complemento_raw):
        numero = normalizar_numero(complemento_raw)
        complemento_raw = ""

    bairro = normalizar_localidade(bairro_raw)
    cidade = normalizar_localidade(cidade_raw)
    uf = normalizar_uf(uf_raw)

    if not uf and cidade_raw and normalizar_uf(cidade_raw):
        uf = normalizar_uf(cidade_raw)
        cidade = ""

    if not uf and bairro_raw and normalizar_uf(bairro_raw):
        uf = normalizar_uf(bairro_raw)
        bairro = ""

    cep = formatar_cep(cep_raw) or cep_embutido

    if not cidade and cep_raw and parece_texto_localidade(cep_raw):
        cidade = normalizar_localidade(cep_raw)

    if not bairro and cep_raw and parece_bairro_valor(cep_raw):
        bairro = normalizar_localidade(cep_raw)

    complemento = limpar_texto(complemento_raw)
    if complemento and normalizar_localidade(complemento) == "":
        if len(normalizar_chave(complemento)) < 2:
            complemento = ""

    campos = {
        "tipo": tipo,
        "rua": rua,
        "numero": numero,
        "bairro": bairro,
        "cidade": cidade,
        "uf": uf,
        "cep": cep,
        "complemento": complemento,
    }

    campos = aplicar_realocacao_inteligente(campos)
    campos = corrigir_campos_localmente(campos)
    return campos


def linha_tem_algum_dado(campos):
    return any(limpar_texto(v) for v in campos.values())


def avaliar_linha(campos) -> Tuple[bool, List[str]]:
    via = montar_logradouro(campos.get("tipo", ""), campos.get("rua", ""))
    motivos = []

    if rua_suspeita(via):
        motivos.append("logradouro_suspeito")

    if bairro_suspeito(campos.get("bairro", "")):
        motivos.append("bairro_suspeito")

    if cidade_suspeita(campos.get("cidade", "")):
        motivos.append("cidade_suspeita")

    uf = campos.get("uf", "")
    if uf and not uf_valida(uf):
        motivos.append("uf_invalida")
    if not uf:
        motivos.append("uf_vazia")

    cep = campos.get("cep", "")
    if cep and not cep_valido(cep):
        motivos.append("cep_invalido")
    if not cep:
        motivos.append("cep_vazio")

    numero = campos.get("numero", "")
    if numero and not numero_valido(numero):
        motivos.append("numero_suspeito")

    suspeita = len(motivos) > 0
    return suspeita, motivos


def consegue_busca_endereco(campos):
    via = montar_logradouro(campos.get("tipo", ""), campos.get("rua", ""))
    cidade = campos.get("cidade", "")
    uf = campos.get("uf", "")
    return (
        len(normalizar_chave(via)) >= 4
        and len(normalizar_chave(cidade)) >= 3
        and uf_valida(uf)
    )


def consegue_fallback_por_cidade(campos):
    return len(normalizar_chave(campos.get("cidade", ""))) >= 3 and uf_valida(campos.get("uf", ""))


def pode_preencher_cep_por_fallback(cep_atual: str) -> bool:
    return not cep_valido(cep_atual)


def corrigir_campos_localmente(campos: Dict[str, str]) -> Dict[str, str]:
    campos["uf"] = normalizar_uf(campos.get("uf", ""))
    campos["cep"] = formatar_cep(campos.get("cep", ""))
    campos["numero"] = normalizar_numero(campos.get("numero", ""))
    campos["bairro"] = normalizar_localidade(campos.get("bairro", ""))
    campos["cidade"] = normalizar_localidade(campos.get("cidade", ""))

    bairro = campos.get("bairro", "")
    cidade = campos.get("cidade", "")

    if bairro and cidade:
        bairro_parece_cidade = parece_cidade_texto(bairro) and not parece_bairro_texto(bairro)
        cidade_parece_bairro = parece_bairro_texto(cidade)

        if bairro_parece_cidade and cidade_parece_bairro:
            campos["bairro"], campos["cidade"] = cidade, bairro

    if (
        not campos.get("cidade")
        and campos.get("bairro")
        and parece_cidade_texto(campos["bairro"])
        and not parece_bairro_texto(campos["bairro"])
    ):
        campos["cidade"] = campos["bairro"]
        campos["bairro"] = ""

    if not campos.get("bairro") and campos.get("cidade") and parece_bairro_texto(campos["cidade"]):
        campos["bairro"] = campos["cidade"]
        campos["cidade"] = ""

    tipo_emb, nome_emb = separar_tipo_nome_logradouro(
        montar_logradouro(campos.get("tipo", ""), campos.get("rua", ""))
    )
    if tipo_emb and nome_emb:
        campos["tipo"] = tipo_emb
        campos["rua"] = corrigir_rua(nome_emb)

    if not campos.get("rua") and campos.get("tipo"):
        tipo_emb2, nome_emb2 = separar_tipo_nome_logradouro(campos["tipo"])
        if tipo_emb2 and nome_emb2:
            campos["tipo"] = tipo_emb2
            campos["rua"] = corrigir_rua(nome_emb2)

    return campos

# =========================================================
# APLICAÇÃO DE RETORNOS
# =========================================================

def aplicar_retorno_correios_local(campos, data):
    if not data:
        return campos

    tipo, nome_logradouro = separar_tipo_nome_logradouro(data.get("logradouro", ""))

    if tipo:
        campos["tipo"] = tipo
    if nome_logradouro:
        campos["rua"] = nome_logradouro
    if data.get("bairro"):
        campos["bairro"] = normalizar_localidade(data.get("bairro", ""))
    if data.get("localidade"):
        campos["cidade"] = normalizar_localidade(data.get("localidade", ""))
    if data.get("uf"):
        campos["uf"] = normalizar_uf(data.get("uf", ""))
    if data.get("cep"):
        campos["cep"] = formatar_cep(data.get("cep", ""))
    if data.get("complemento") and not campos.get("complemento"):
        campos["complemento"] = limpar_texto(data.get("complemento", ""))

    campos = aplicar_realocacao_inteligente(campos)
    return corrigir_campos_localmente(campos)


def aplicar_retorno_viacep(campos, data):
    if not data:
        return campos

    tipo, nome_logradouro = separar_tipo_nome_logradouro(data.get("logradouro", ""))

    if tipo:
        campos["tipo"] = tipo
    if nome_logradouro:
        campos["rua"] = nome_logradouro
    if data.get("bairro"):
        campos["bairro"] = normalizar_localidade(data.get("bairro", ""))
    if data.get("localidade"):
        campos["cidade"] = normalizar_localidade(data.get("localidade", ""))
    if data.get("uf"):
        campos["uf"] = normalizar_uf(data.get("uf", ""))
    if data.get("cep"):
        campos["cep"] = formatar_cep(data.get("cep", ""))

    campos = aplicar_realocacao_inteligente(campos)
    return corrigir_campos_localmente(campos)


def aplicar_retorno_osm(campos, item):
    if not item:
        return campos

    road = pegar_osm_address(item, "road", "pedestrian", "residential", "footway", "path")
    house_number = pegar_osm_address(item, "house_number")
    suburb = pegar_osm_address(item, "suburb", "neighbourhood")
    city = pegar_osm_address(item, "city", "town", "municipality", "village")
    state = pegar_osm_address(item, "state")
    postcode = pegar_osm_address(item, "postcode")

    tipo, nome = separar_tipo_nome_logradouro(road)

    if tipo:
        campos["tipo"] = tipo
    if nome:
        campos["rua"] = nome
    if house_number and not campos.get("numero"):
        campos["numero"] = normalizar_numero(house_number)
    if suburb:
        campos["bairro"] = normalizar_localidade(suburb)
    if city:
        campos["cidade"] = normalizar_localidade(city)
    if state and not campos.get("uf"):
        campos["uf"] = normalizar_uf(state)
    if postcode:
        campos["cep"] = formatar_cep(postcode) or campos.get("cep", "")

    campos = aplicar_realocacao_inteligente(campos)
    return corrigir_campos_localmente(campos)


def aplicar_fallback_cep_cidade(campos, data):
    if not data:
        return campos

    if pode_preencher_cep_por_fallback(campos.get("cep", "")) and data.get("cep"):
        campos["cep"] = formatar_cep(data.get("cep", ""))

    if not campos.get("cidade") and data.get("localidade"):
        campos["cidade"] = normalizar_localidade(data.get("localidade", ""))

    if not campos.get("uf") and data.get("uf"):
        campos["uf"] = normalizar_uf(data.get("uf", ""))

    campos = aplicar_realocacao_inteligente(campos)
    return corrigir_campos_localmente(campos)

# =========================================================
# SOBRESCRITA DAS COLUNAS ORIGINAIS
# =========================================================

def sobrescrever_colunas_originais(ws, row_idx, cols_map, resultado, uppercase_addresses=False):
    campos = resultado["campos"]

    def set_if_has_value(chave, valor):
        col = cols_map.get(chave)
        if not col:
            return

        if chave in {"tipo", "rua", "bairro", "cidade", "complemento"}:
            valor = texto_saida_endereco(valor, uppercase_addresses)
        else:
            valor = limpar_texto(valor)

        if valor:
            ws.cell(row=row_idx, column=col, value=valor)

    if cols_map.get("tipo"):
        tipo_final = texto_saida_endereco(campos.get("tipo", ""), uppercase_addresses)
        rua_final = texto_saida_endereco(campos.get("rua", ""), uppercase_addresses)

        if tipo_final:
            ws.cell(row=row_idx, column=cols_map["tipo"], value=tipo_final)
        if cols_map.get("rua") and rua_final:
            ws.cell(row=row_idx, column=cols_map["rua"], value=rua_final)
    else:
        if cols_map.get("rua"):
            logradouro_novo = montar_logradouro(campos.get("tipo", ""), campos.get("rua", ""))
            logradouro_novo = texto_saida_endereco(logradouro_novo, uppercase_addresses)
            if logradouro_novo:
                ws.cell(row=row_idx, column=cols_map["rua"], value=logradouro_novo)

    if numero_valido(campos.get("numero", "")):
        set_if_has_value("numero", normalizar_numero(campos.get("numero", "")))

    bairro_col = cols_map.get("bairro")
    if bairro_col:
        bairro_original = limpar_texto(ws.cell(row=row_idx, column=bairro_col).value)
        bairro_final = limpar_texto(campos.get("bairro", ""))
        if bairro_final and not bairro_suspeito(bairro_final):
            ws.cell(
                row=row_idx,
                column=bairro_col,
                value=texto_saida_endereco(bairro_final, uppercase_addresses)
            )
        elif bairro_original_invalido_e_nao_aproveitavel(bairro_original):
            ws.cell(row=row_idx, column=bairro_col, value="")

    cidade_col = cols_map.get("cidade")
    if cidade_col:
        cidade_original = limpar_texto(ws.cell(row=row_idx, column=cidade_col).value)
        cidade_final = limpar_texto(campos.get("cidade", ""))
        if cidade_final and not cidade_suspeita(cidade_final):
            ws.cell(
                row=row_idx,
                column=cidade_col,
                value=texto_saida_endereco(cidade_final, uppercase_addresses)
            )
        elif cidade_original_invalida_e_nao_aproveitavel(cidade_original):
            ws.cell(row=row_idx, column=cidade_col, value="")

    uf_col = cols_map.get("uf")
    if uf_col:
        uf_original = limpar_texto(ws.cell(row=row_idx, column=uf_col).value)
        uf_final = normalizar_uf(campos.get("uf", ""))
        if uf_final:
            ws.cell(row=row_idx, column=uf_col, value=uf_final)
        elif uf_original_invalida_e_nao_aproveitavel(uf_original):
            ws.cell(row=row_idx, column=uf_col, value="")

    cep_col = cols_map.get("cep")
    if cep_col:
        cep_original = limpar_texto(ws.cell(row=row_idx, column=cep_col).value)
        cep_final = formatar_cep(campos.get("cep", ""))

        if cep_final:
            ws.cell(row=row_idx, column=cep_col, value=cep_final)
        else:
            if cep_original_invalido_e_nao_aproveitavel(cep_original):
                ws.cell(row=row_idx, column=cep_col, value="")

    if limpar_texto(campos.get("complemento", "")):
        set_if_has_value("complemento", campos.get("complemento", ""))

# =========================================================
# PROCESSAMENTO DA LINHA
# =========================================================

def processar_linha(campos, ctx, mode):
    suspeita_antes, _ = avaliar_linha(campos)

    usou_dne_cep = False
    usou_dne_endereco = False
    usou_dne_cidade = False
    usou_viacep_cep = False
    usou_viacep_endereco = False
    usou_osm = False
    osm_item = None

    if cep_valido(campos.get("cep", "")):
        data_local_cep = correios_por_cep_local(campos.get("cep", ""), ctx)
        if data_local_cep:
            campos = aplicar_retorno_correios_local(campos, data_local_cep)
            usou_dne_cep = True

    if not usou_dne_cep and consegue_busca_endereco(campos):
        via = montar_logradouro(campos.get("tipo", ""), campos.get("rua", ""))
        data_local_end = correios_por_endereco_local(
            campos.get("uf", ""),
            campos.get("cidade", ""),
            via,
            campos.get("bairro", ""),
            ctx,
        )
        if data_local_end:
            campos = aplicar_retorno_correios_local(campos, data_local_end)
            usou_dne_endereco = True

    if (
        not usou_dne_cep
        and not usou_dne_endereco
        and pode_preencher_cep_por_fallback(campos.get("cep", ""))
        and consegue_fallback_por_cidade(campos)
    ):
        data_cidade = correios_aproximado_por_cidade_local(
            campos.get("uf", ""),
            campos.get("cidade", ""),
            ctx,
        )
        if data_cidade:
            campos = aplicar_fallback_cep_cidade(campos, data_cidade)
            usou_dne_cidade = True

    if mode == "completo":
        if cep_valido(campos.get("cep", "")) and not usou_dne_cep:
            data_cep = viacep_por_cep(campos.get("cep", ""), ctx)
            if data_cep:
                campos = aplicar_retorno_viacep(campos, data_cep)
                usou_viacep_cep = True

        if not usou_dne_endereco and not usou_viacep_cep and consegue_busca_endereco(campos):
            via = montar_logradouro(campos.get("tipo", ""), campos.get("rua", ""))
            data_end = viacep_por_endereco(
                campos.get("uf", ""),
                campos.get("cidade", ""),
                via,
                campos.get("bairro", ""),
                ctx,
            )
            if data_end:
                campos = aplicar_retorno_viacep(campos, data_end)
                usou_viacep_endereco = True

    endereco_consultado = montar_endereco_consulta(campos)

    if mode == "completo" and len(normalizar_chave(endereco_consultado)) >= 8:
        osm_item = nominatim_buscar(endereco_consultado, ctx)
        if osm_item:
            campos = aplicar_retorno_osm(campos, osm_item)
            usou_osm = True

    campos = aplicar_realocacao_inteligente(campos)
    campos = corrigir_campos_localmente(campos)

    return {
        "campos": campos,
        "meta": {
            "suspeita_antes": suspeita_antes,
            "usou_dne_cep": usou_dne_cep,
            "usou_dne_endereco": usou_dne_endereco,
            "usou_dne_cidade": usou_dne_cidade,
            "usou_viacep_cep": usou_viacep_cep,
            "usou_viacep_endereco": usou_viacep_endereco,
            "usou_osm": usou_osm,
            "osm_item": osm_item,
        },
    }

# =========================================================
# WORKBOOK
# =========================================================

def contar_total_linhas_processaveis(wb):
    total = 0

    for ws in wb.worksheets:
        if ws.max_row < 1:
            continue

        det = detectar_mapeamento_ws(ws)
        if not det["source"]:
            continue

        start_row = det["start_row"]
        cols_map = det["cols"]

        if start_row > ws.max_row:
            continue

        for row_idx in range(start_row, ws.max_row + 1):
            campos = ler_campos_da_linha(ws, row_idx, cols_map)
            if linha_tem_algum_dado(campos):
                total += 1

    return total


def processar_workbook_bytes(input_bytes: bytes, mode: str, ext: str, job_id=None, uppercase_addresses=False) -> bytes:
    ctx = ConsultaContexto()

    try:
        keep_vba = ext == ".xlsm"
        wb = load_workbook(BytesIO(input_bytes), keep_vba=keep_vba)

        limpar_artefatos_antigos(wb)

        total = contar_total_linhas_processaveis(wb)

        if job_id:
            update_job(
                job_id,
                total=total,
                current=0,
                progress=0,
                status="processando",
                message="Lendo planilha..."
            )

        current = 0

        for ws in wb.worksheets:
            if ws.max_row < 1:
                continue

            det = detectar_mapeamento_ws(ws)
            if not det["source"]:
                continue

            cols_map = det["cols"]
            start_row = det["start_row"]

            if start_row > ws.max_row:
                continue

            for row_idx in range(start_row, ws.max_row + 1):
                campos = ler_campos_da_linha(ws, row_idx, cols_map)

                if not linha_tem_algum_dado(campos):
                    continue

                resultado = processar_linha(campos, ctx, mode)
                sobrescrever_colunas_originais(
                    ws,
                    row_idx,
                    cols_map,
                    resultado,
                    uppercase_addresses=uppercase_addresses
                )

                current += 1

                if job_id and total > 0:
                    pct = int((current / total) * 100)
                    update_job(
                        job_id,
                        current=current,
                        total=total,
                        progress=min(pct, 100),
                        status="processando",
                        message=f"Modo {mode}: {current}/{total} • Aba {ws.title}"
                    )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()
    finally:
        ctx.close()


def executar_job(job_id: str, input_bytes: bytes, ext: str, nome_arquivo: str, mode: str, uppercase_addresses: bool = False):
    try:
        update_job(
            job_id,
            status="processando",
            message=f"Iniciando processamento no modo {mode}..."
        )

        output_bytes = processar_workbook_bytes(
            input_bytes=input_bytes,
            mode=mode,
            ext=ext,
            job_id=job_id,
            uppercase_addresses=uppercase_addresses
        )

        output_ext = ".xlsm" if ext == ".xlsm" else ".xlsx"

        update_job(
            job_id,
            status="concluido",
            progress=100,
            done=True,
            output_bytes=output_bytes,
            output_name=f"geopadroniza_{mode}_{Path(nome_arquivo).stem}{output_ext}",
            message="Processamento concluído."
        )
    except Exception as e:
        update_job(
            job_id,
            status="erro",
            done=True,
            error=str(e),
            message=f"Erro: {str(e)}"
        )

# =========================================================
# ROTAS
# =========================================================

@app.get("/", response_class=HTMLResponse)
def home():
    html_file = BASE_DIR / "index.html"
    if html_file.exists():
        return HTMLResponse(html_file.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>GeoPadroniza</h1><p>Coloque o index.html na mesma pasta do app.py.</p>")


@app.get("/logo.jpg")
def logo_jpg():
    logo_file = BASE_DIR / "logo.jpg"
    if not logo_file.exists():
        raise HTTPException(status_code=404, detail="logo.jpg não encontrado na pasta do app.")
    return FileResponse(logo_file, media_type="image/jpeg")


@app.get("/favicon.jpg")
def favicon_jpg():
    favicon_file = BASE_DIR / "favicon.jpg"
    if not favicon_file.exists():
        raise HTTPException(status_code=404, detail="favicon.jpg não encontrado na pasta do app.")
    return FileResponse(favicon_file, media_type="image/jpeg")


@app.get("/favicon.ico")
def favicon_ico():
    favicon_file = BASE_DIR / "favicon.jpg"
    if favicon_file.exists():
        return FileResponse(favicon_file, media_type="image/jpeg")
    return Response(status_code=204)


@app.get("/health")
def health():
    return {
        "ok": True,
        "app": APP_NAME,
        "version": APP_VERSION,
        "base_local_correios_ativa": Path(DNE_DB_PATH).exists(),
        "base_local_correios_path": DNE_DB_PATH,
        "sobrescreve_colunas_originais": True,
        "cria_colunas_geo": False,
        "cria_aba_log": False,
        "remove_geo_antigo_se_existir": True,
        "usa_jknor_fixo": False,
        "fallback_cep_aproximado_por_cidade": True,
        "limpa_cep_invalido_quando_nao_recupera": True,
        "realoca_valor_para_coluna_certa": True,
        "suporta_enderecos_caixa_alta": True,
        "modos": ["ultra", "rapido", "completo"],
        "extensoes_seguras": [".xlsx", ".xlsm"],
    }


@app.post("/process")
async def process(
    file: UploadFile = File(...),
    mode: str = Form("ultra"),
    uppercase_addresses: str = Form("false")
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Arquivo inválido.")

    ext = Path(file.filename).suffix.lower()
    if ext not in [".xlsx", ".xlsm"]:
        raise HTTPException(
            status_code=400,
            detail="Para preservar a estrutura, use arquivo .xlsx ou .xlsm."
        )

    mode = (mode or "ultra").strip().lower()
    if mode not in MODES:
        raise HTTPException(status_code=400, detail="Modo inválido. Use 'ultra', 'rapido' ou 'completo'.")

    uppercase_addresses_bool = form_bool(uppercase_addresses)

    input_bytes = await file.read()
    if not input_bytes:
        raise HTTPException(status_code=400, detail="Arquivo vazio.")

    job_id = init_job(file.filename, mode)
    JOBS[job_id]["uppercase_addresses"] = uppercase_addresses_bool

    thread = threading.Thread(
        target=executar_job,
        args=(job_id, input_bytes, ext, file.filename, mode, uppercase_addresses_bool),
        daemon=True
    )
    thread.start()

    return {
        "ok": True,
        "job_id": job_id,
        "mode": mode,
        "uppercase_addresses": uppercase_addresses_bool
    }


@app.get("/status/{job_id}")
def status_job(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job não encontrado.")

    return {
        "id": job["id"],
        "mode": job["mode"],
        "status": job["status"],
        "progress": job["progress"],
        "current": job["current"],
        "total": job["total"],
        "message": job["message"],
        "done": job["done"],
        "error": job["error"],
        "uppercase_addresses": job.get("uppercase_addresses", False),
    }


@app.get("/download/{job_id}")
def download_job(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job não encontrado.")

    if not job["done"]:
        raise HTTPException(status_code=400, detail="O processamento ainda não terminou.")

    if job["error"]:
        raise HTTPException(status_code=500, detail=job["error"])

    if not job["output_bytes"]:
        raise HTTPException(status_code=500, detail="Arquivo final não encontrado.")

    headers = {
        "Content-Disposition": f'attachment; filename="{job["output_name"]}"'
    }

    return StreamingResponse(
        BytesIO(job["output_bytes"]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )